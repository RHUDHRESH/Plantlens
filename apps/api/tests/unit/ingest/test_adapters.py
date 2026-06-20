"""Unit tests for offline ingestion CSV/XLSX adapters."""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO

import pytest
from openpyxl import Workbook
from pydantic import ValidationError

from app.ingest.adapters import (
    AdapterParseError,
    AdapterResult,
    CsvAdapter,
    XlsxAdapter,
    clean_cell,
    clean_header,
)
from app.schemas.ingest.artifact import RawArtifact

RUN_ID = "run_00000000-0000-4000-8000-000000000001"
ART_ID = "art_00000000-0000-4000-8000-000000000001"
SHA256 = "a" * 64
TS = datetime(2026, 6, 20, 12, 0, 0, tzinfo=UTC)


def _artifact(**overrides: object) -> RawArtifact:
    base = {
        "artifact_id": ART_ID,
        "run_id": RUN_ID,
        "received_at_utc": TS,
        "size_bytes": 0,
        "sha256": SHA256,
        "source_channel": "upload",
        "raw_uri": "memory://raw/test",
        "document_kind": None,
        "detection_confidence": 0.0,
        "detection_signals": [],
    }
    base.update(overrides)
    return RawArtifact(**base)  # type: ignore[arg-type]


def test_clean_header_normalizes_common_noise():
    assert clean_header("Asset Label") == "asset_label"
    assert clean_header(" signal-label ") == "signal_label"
    assert clean_header("Tag/Hint") == "tag_hint"
    assert clean_header("  AC Output  Voltage ") == "ac_output_voltage"


def test_clean_cell_preserves_values_and_blanks():
    assert clean_cell(None) is None
    assert clean_cell("") is None
    assert clean_cell("   ") is None
    assert clean_cell(" V ") == "V"
    assert clean_cell(24) == "24"


def test_csv_adapter_extracts_records_with_source_refs():
    csv_bytes = (
        b"asset_label,signal_label,tag_hint,unit,side\n"
        b"Solar Charger,Output Voltage,V1,V,dc\n"
        b"Mains Charger,Output Current,I2,A,dc\n"
    )
    result = CsvAdapter().extract_records(artifact=_artifact(), content=csv_bytes)

    assert len(result.records) == 2
    assert result.records[0].row_number == 2
    assert result.records[1].row_number == 3
    assert result.records[0].sheet_name is None
    assert result.records[0].fields["asset_label"] == "Solar Charger"
    assert result.records[0].fields["signal_label"] == "Output Voltage"
    assert result.records[0].source_ref.artifact_id == ART_ID
    assert result.records[0].source_ref.artifact_sha256 == SHA256
    assert result.metadata["record_count"] == 2


def test_csv_adapter_uses_utf8_sig():
    csv_bytes = (
        b"\xef\xbb\xbfasset_label,signal_label\n"
        b"Solar Charger,Output Voltage\n"
    )
    result = CsvAdapter().extract_records(artifact=_artifact(), content=csv_bytes)

    assert result.metadata["cleaned_headers"] == ["asset_label", "signal_label"]
    assert "asset_label" in result.records[0].fields


def test_csv_adapter_skips_empty_rows_but_preserves_row_numbers():
    csv_bytes = (
        b"asset_label,signal_label\n"
        b"Solar Charger,Output Voltage\n"
        b",\n"
        b"Mains Charger,Output Current\n"
    )
    result = CsvAdapter().extract_records(artifact=_artifact(), content=csv_bytes)

    assert len(result.records) == 2
    assert result.records[0].row_number == 2
    assert result.records[1].row_number == 4


def test_csv_adapter_rejects_duplicate_headers_after_cleaning():
    csv_bytes = b"Asset Label,asset-label\nSolar Charger,Output Voltage\n"
    with pytest.raises(AdapterParseError):
        CsvAdapter().extract_records(artifact=_artifact(), content=csv_bytes)


def test_csv_adapter_warns_for_missing_columns():
    csv_bytes = (
        b"asset_label,signal_label,tag_hint,unit,side\n"
        b"Solar Charger,Output Voltage,V1\n"
    )
    result = CsvAdapter().extract_records(artifact=_artifact(), content=csv_bytes)

    assert "row_has_missing_columns" in result.records[0].parser_warnings
    assert result.records[0].fields["unit"] is None
    assert result.records[0].fields["side"] is None


def test_csv_adapter_warns_for_extra_columns():
    csv_bytes = (
        b"asset_label,signal_label\n"
        b"Solar Charger,Output Voltage,extra\n"
    )
    result = CsvAdapter().extract_records(artifact=_artifact(), content=csv_bytes)

    assert "row_has_extra_columns" in result.records[0].parser_warnings
    assert set(result.records[0].fields.keys()) == {"asset_label", "signal_label"}


def test_xlsx_adapter_extracts_records_with_sheet_names():
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Signals"
    sheet.append(["asset_label", "signal_label", "tag_hint", "unit", "side"])
    sheet.append(["Solar Charger", "Output Voltage", "V1", "V", "dc"])
    buffer = BytesIO()
    workbook.save(buffer)

    result = XlsxAdapter().extract_records(artifact=_artifact(), content=buffer.getvalue())

    assert len(result.records) == 1
    assert result.records[0].sheet_name == "Signals"
    assert result.records[0].row_number == 2
    assert result.records[0].fields["asset_label"] == "Solar Charger"
    assert result.records[0].source_ref.sheet_name == "Signals"


def test_xlsx_adapter_skips_empty_sheet_with_warning():
    workbook = Workbook()
    empty_sheet = workbook.active
    assert empty_sheet is not None
    empty_sheet.title = "Empty"
    workbook.create_sheet("Signals")
    signals = workbook["Signals"]
    signals.append(["asset_label", "signal_label"])
    signals.append(["Solar Charger", "Output Voltage"])
    buffer = BytesIO()
    workbook.save(buffer)

    result = XlsxAdapter().extract_records(artifact=_artifact(), content=buffer.getvalue())

    assert any("sheet_empty:Empty" in warning for warning in result.warnings)
    assert len(result.records) == 1


def test_xlsx_adapter_rejects_duplicate_headers_after_cleaning():
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["Asset Label", "asset-label"])
    sheet.append(["Solar Charger", "Output Voltage"])
    buffer = BytesIO()
    workbook.save(buffer)

    with pytest.raises(AdapterParseError):
        XlsxAdapter().extract_records(artifact=_artifact(), content=buffer.getvalue())


def test_xlsx_adapter_metadata_contains_sheet_details():
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Signals"
    sheet.append(["asset_label", "signal_label"])
    sheet.append(["Solar Charger", "Output Voltage"])
    buffer = BytesIO()
    workbook.save(buffer)

    result = XlsxAdapter().extract_records(artifact=_artifact(), content=buffer.getvalue())

    assert "Signals" in result.metadata["sheet_names"]
    assert result.metadata["header_rows"]["Signals"] == 1
    assert result.metadata["cleaned_headers"]["Signals"] == ["asset_label", "signal_label"]
    assert result.metadata["record_count"] == 1


def test_adapters_do_not_mutate_artifact():
    artifact = _artifact(document_kind="signal_list", detection_confidence=0.8)
    csv_bytes = b"asset_label,signal_label\nSolar Charger,Output Voltage\n"
    result = CsvAdapter().extract_records(artifact=artifact, content=csv_bytes)

    assert result.artifact is artifact
    assert artifact.document_kind == "signal_list"
    assert artifact.detection_confidence == 0.8
    assert artifact.detection_signals == []


def test_adapter_result_rejects_wrong_shape_if_pydantic_used():
    artifact = _artifact()
    with pytest.raises(ValidationError):
        AdapterResult(
            artifact=artifact,
            records=[],
            warnings=[],
            metadata={},
            unexpected_field="nope",  # type: ignore[call-arg]
        )