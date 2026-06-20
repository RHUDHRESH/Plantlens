"""XLSX tabular adapter for offline authored-knowledge ingestion."""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from app.schemas.ingest.artifact import RawArtifact
from app.schemas.ingest.record import RawRecord, SourceRef

from app.ingest.adapters.base import (
    AdapterResult,
    clean_cell,
    clean_header,
    make_raw_id,
    reject_duplicate_headers,
    row_is_empty,
)


class XlsxAdapter:
    """Extract RawRecord rows from XLSX bytes without semantic interpretation."""

    def extract_records(self, *, artifact: RawArtifact, content: bytes) -> AdapterResult:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)

        records: list[RawRecord] = []
        warnings: list[str] = []
        sheet_names: list[str] = []
        header_rows: dict[str, int | None] = {}
        cleaned_headers_by_sheet: dict[str, list[str]] = {}
        skipped_empty_rows = 0

        for worksheet in workbook.worksheets:
            if worksheet.sheet_state != "visible":
                warnings.append(f"sheet_hidden_skipped:{worksheet.title}")
                continue

            sheet_names.append(worksheet.title)
            sheet_records, sheet_warnings, sheet_skipped, header_row, cleaned_headers = (
                self._extract_sheet(worksheet, artifact)
            )
            records.extend(sheet_records)
            warnings.extend(sheet_warnings)
            skipped_empty_rows += sheet_skipped
            header_rows[worksheet.title] = header_row
            cleaned_headers_by_sheet[worksheet.title] = cleaned_headers

        workbook.close()

        metadata: dict[str, Any] = {
            "sheet_names": sheet_names,
            "sheet_count": len(sheet_names),
            "header_rows": header_rows,
            "cleaned_headers": cleaned_headers_by_sheet,
            "skipped_empty_rows": skipped_empty_rows,
            "record_count": len(records),
        }
        return AdapterResult(
            artifact=artifact,
            records=records,
            warnings=warnings,
            metadata=metadata,
        )

    def _extract_sheet(
        self,
        worksheet: Worksheet,
        artifact: RawArtifact,
    ) -> tuple[list[RawRecord], list[str], int, int | None, list[str]]:
        records: list[RawRecord] = []
        warnings: list[str] = []
        skipped_empty_rows = 0
        header_row_number: int | None = None
        cleaned_headers: list[str] = []
        original_headers: list[str] = []

        for row_number, row in enumerate(
            worksheet.iter_rows(values_only=True),
            start=1,
        ):
            values = list(row)
            if row_is_empty(values):
                if header_row_number is None:
                    skipped_empty_rows += 1
                    continue
                skipped_empty_rows += 1
                continue

            if header_row_number is None:
                header_row_number = row_number
                original_headers = [str(value) if value is not None else "" for value in values]
                cleaned_headers = [clean_header(header) for header in original_headers]
                reject_duplicate_headers(
                    cleaned_headers,
                    context=f"xlsx sheet {worksheet.title}",
                )
                continue

            row_warnings = self._width_warnings(values, cleaned_headers)
            fields = self._build_fields(values, cleaned_headers)
            if not fields:
                skipped_empty_rows += 1
                continue

            records.append(
                RawRecord(
                    raw_id=make_raw_id(),
                    artifact_id=artifact.artifact_id,
                    run_id=artifact.run_id,
                    row_number=row_number,
                    sheet_name=worksheet.title,
                    fields=fields,
                    source_ref=SourceRef(
                        artifact_id=artifact.artifact_id,
                        artifact_sha256=artifact.sha256,
                        row_number=row_number,
                        sheet_name=worksheet.title,
                        column_name=None,
                        cell_ref=None,
                    ),
                    extracted_at_utc=datetime.now(UTC),
                    parser_warnings=row_warnings,
                )
            )

        if header_row_number is None:
            warnings.append(f"sheet_empty:{worksheet.title}")

        return records, warnings, skipped_empty_rows, header_row_number, cleaned_headers

    def _width_warnings(self, row: list[object], cleaned_headers: list[str]) -> list[str]:
        warnings: list[str] = []
        if len(row) > len(cleaned_headers):
            warnings.append("row_has_extra_columns")
        elif len(row) < len(cleaned_headers):
            warnings.append("row_has_missing_columns")
        return warnings

    def _build_fields(self, row: list[object], cleaned_headers: list[str]) -> dict[str, str | None]:
        fields: dict[str, str | None] = {}
        for index, header in enumerate(cleaned_headers):
            value = row[index] if index < len(row) else None
            fields[header] = clean_cell(value)
        return fields